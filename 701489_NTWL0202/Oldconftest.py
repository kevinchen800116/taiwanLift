import pytest
from openpyxl import load_workbook

workbook = load_workbook("NTWL0202Data.xlsx")
sheet = workbook["personinfoAll"]
sheet1 = workbook["SaleReportAll"]
sheet2 = workbook["productAll"]
sheet3 = workbook["BenefitInfo"]
sheet4 =workbook["Agent"]


personinfoAll={}
for col in range(1,sheet.max_column+1):
        personinfoAll[sheet.cell(3, col).value] = sheet.cell(4, col).value




SaleReportAll={}
for col in range(1,sheet1.max_column+1):
        SaleReportAll[sheet1.cell(3, col).value] = sheet1.cell(4, col).value


productAll={}
for col in range(1,sheet2.max_column+1):
        productAll[sheet2.cell(3, col).value] = sheet2.cell(4, col).value


BenefitInfo={}
for col in range(1,sheet3.max_column+1):
        BenefitInfo[sheet3.cell(3, col).value] = sheet3.cell(4, col).value


Agent={}
for col in range(1,sheet4.max_column+1):
        Agent[sheet4.cell(3, col).value] = sheet4.cell(4, col).value



# ------------------------------------------------------------------------
# data = [(1, 2, 3), (4, 5, 9)]
# All_test_Data=[(personinfoAll,BenefitInfo,SaleReportAll,productAll),(personinfo1,BenefitInfo1,SaleReport1,product1)]
All_test_Data=[(personinfoAll,BenefitInfo,SaleReportAll,productAll)]

@pytest.fixture(scope="session",params=All_test_Data)
def my_fixture(request):
    return request.param


