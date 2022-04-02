import pytest
from openpyxl import load_workbook

workbook = load_workbook("NITU0905_02_Data.xlsx")
sheet = workbook["personinfoAll"]
sheet1 = workbook["SaleReportAll"]
sheet2 = workbook["productAll"]
sheet3 = workbook["BenefitInfo"]
sheet4 =workbook["Agent"]


personinfoAll={}
personinfo1={}
personinfo2={}
personinfo3={}
personinfo4={}

for col in range(1,sheet.max_column+1):
        personinfoAll[sheet.cell(3, col).value] = sheet.cell(4, col).value
        personinfo1[sheet.cell(3, col).value] = sheet.cell(5, col).value
        personinfo2[sheet.cell(3, col).value] = sheet.cell(6, col).value
        personinfo3[sheet.cell(3, col).value] = sheet.cell(7, col).value
        personinfo4[sheet.cell(3, col).value] = sheet.cell(8, col).value

print(personinfoAll)
# print(personinfo1)


SaleReportAll={}
SaleReport1={}
SaleReport2={}
SaleReport3={}
SaleReport4={}

for col in range(1,sheet1.max_column+1):
        SaleReportAll[sheet1.cell(3, col).value] = sheet1.cell(4, col).value
        SaleReport1[sheet1.cell(3, col).value] = sheet1.cell(5, col).value
        SaleReport2[sheet1.cell(3, col).value] = sheet1.cell(6, col).value
        SaleReport3[sheet1.cell(3, col).value] = sheet1.cell(7, col).value
        SaleReport4[sheet1.cell(3, col).value] = sheet1.cell(8, col).value

print(SaleReportAll)
# print(SaleReport1)

productAll={}
product1={}
product2={}
product3={}
product4={}

for col in range(1,sheet2.max_column+1):
        productAll[sheet2.cell(3, col).value] = sheet2.cell(4, col).value
        product1[sheet2.cell(3, col).value] = sheet2.cell(5, col).value
        product2[sheet2.cell(3, col).value] = sheet2.cell(6, col).value
        product3[sheet2.cell(3, col).value] = sheet2.cell(7, col).value
        product4[sheet2.cell(3, col).value] = sheet2.cell(8, col).value

print(productAll)

BenefitInfo={}
BenefitInfo1={}
BenefitInfo2={}
BenefitInfo3={}
BenefitInfo4={}

for col in range(1,sheet3.max_column+1):
        BenefitInfo[sheet3.cell(3, col).value] = sheet3.cell(4, col).value
        BenefitInfo1[sheet3.cell(3, col).value] = sheet3.cell(5, col).value
        BenefitInfo2[sheet3.cell(3, col).value] = sheet3.cell(6, col).value
        BenefitInfo3[sheet3.cell(3, col).value] = sheet3.cell(7, col).value
        BenefitInfo4[sheet3.cell(3, col).value] = sheet3.cell(8, col).value
print(BenefitInfo)

Agent={}
for col in range(1,sheet4.max_column+1):
        Agent[sheet4.cell(3, col).value] = sheet4.cell(4, col).value
print(Agent)



# ------------------------------------------------------------------------
# data = [(1, 2, 3), (4, 5, 9)]
# All_test_Data=[(personinfoAll,BenefitInfo,SaleReportAll,productAll),(personinfo1,BenefitInfo1,SaleReport1,product1)]
All_test_Data=[
        (personinfoAll,BenefitInfo,SaleReportAll,productAll),
        (personinfo1,BenefitInfo1,SaleReport1,product1),
        (personinfo2,BenefitInfo2,SaleReport1,product2),
        (personinfo3,BenefitInfo3,SaleReport1,product3),
        (personinfo4,BenefitInfo4,SaleReport1,product4)
        ]

@pytest.fixture(scope="session",params=All_test_Data)
def my_fixture(request):
    return request.param


