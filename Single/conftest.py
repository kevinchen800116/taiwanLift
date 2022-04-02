import pytest
from openpyxl import load_workbook

workbook = load_workbook("NUIW0202_Data.xlsx")
sheet = workbook["personinfoAll"]
sheet1 = workbook["SaleReportAll"]
sheet2 = workbook["productAll"]
sheet3 = workbook["BenefitInfo"]
sheet4 =workbook["Agent"]

### personinfoAll
test=[]
for row in range(4,sheet.max_row+1):
        ## 定義一個空字典
        dict={}
        for col in range(1,sheet.max_column+1):
                ## 迴圈將sheet內的資料一行一行寫入字典
                dict[sheet.cell(3, col).value] = sheet.cell(row, col).value
        ## 將寫入資料後的字典，加進test串列
        test.append(dict)

### SaleReportAll
test1=[]
for row in range(4,sheet1.max_row+1):
        dict={}
        for col in range(1,sheet1.max_column+1):
                dict[sheet1.cell(3, col).value] = sheet1.cell(row, col).value
        test1.append(dict)

### productAll
test2=[]
for row in range(4,sheet2.max_row+1):
        dict={}
        for col in range(1,sheet2.max_column+1):
                dict[sheet2.cell(3, col).value] = sheet2.cell(row, col).value
        test2.append(dict)

### BenefitInfo
test3=[]
for row in range(4,sheet3.max_row+1):
        dict={}
        for col in range(1,sheet3.max_column+1):
                dict[sheet3.cell(3, col).value] = sheet3.cell(row, col).value
        test3.append(dict)

### Agent
test4=[]
for row in range(4,sheet4.max_row+1):
        dict={}
        for col in range(1,sheet4.max_column+1):
                dict[sheet4.cell(3, col).value] = sheet4.cell(row, col).value
        test4.append(dict)


All_test_Data=[]
# 此處的range為控制每個頁籤內，row的列數。此處選擇test的頁籤內的全部資料，當作最大長度。
# ex:若agent頁籤的資料少於6筆資料，但其他頁籤有6筆資料，則test4會報錯。 
for i in range(0,len(test)):
        All=(test[i],test1[i],test2[i],test3[i],test4[i])
        # print("我是ALL："+str(All))
        All_test_Data.append(All)
        
print("全部資料："+str(All_test_Data))

@pytest.fixture(scope="session",params=All_test_Data)
def my_fixture(request):
    return request.param
    
# personinfoAll={}
# personinfo1={}
# personinfo2={}
# personinfo3={}
# personinfo4={}
# personinfo5={}
# # test=[]
# # for i in range(1,10):
# #         test.append("personinfo"+i)
# # print(test)
# for col in range(1,sheet.max_column+1):
#         personinfoAll[sheet.cell(3, col).value] = sheet.cell(4, col).value
#         personinfo1[sheet.cell(3, col).value] = sheet.cell(5, col).value
#         personinfo2[sheet.cell(3, col).value] = sheet.cell(6, col).value
#         personinfo3[sheet.cell(3, col).value] = sheet.cell(7, col).value
#         personinfo4[sheet.cell(3, col).value] = sheet.cell(8, col).value
#         personinfo5[sheet.cell(3, col).value] = sheet.cell(9, col).value

# print(personinfoAll)
# # print(personinfo1)


# SaleReportAll={}
# SaleReport1={}
# SaleReport2={}
# SaleReport3={}
# SaleReport4={}
# SaleReport5={}

# for col in range(1,sheet1.max_column+1):
#         SaleReportAll[sheet1.cell(3, col).value] = sheet1.cell(4, col).value
#         SaleReport1[sheet1.cell(3, col).value] = sheet1.cell(5, col).value
#         SaleReport2[sheet1.cell(3, col).value] = sheet1.cell(6, col).value
#         SaleReport3[sheet1.cell(3, col).value] = sheet1.cell(7, col).value
#         SaleReport4[sheet1.cell(3, col).value] = sheet1.cell(8, col).value
#         SaleReport5[sheet1.cell(3, col).value] = sheet1.cell(9, col).value

# print(SaleReportAll)
# # print(SaleReport1)

# productAll={}
# product1={}
# product2={}
# product3={}
# product4={}
# product5={}

# for col in range(1,sheet2.max_column+1):
#         productAll[sheet2.cell(3, col).value] = sheet2.cell(4, col).value
#         product1[sheet2.cell(3, col).value] = sheet2.cell(5, col).value
#         product2[sheet2.cell(3, col).value] = sheet2.cell(6, col).value
#         product3[sheet2.cell(3, col).value] = sheet2.cell(7, col).value
#         product4[sheet2.cell(3, col).value] = sheet2.cell(8, col).value
#         product5[sheet2.cell(3, col).value] = sheet2.cell(9, col).value

# print(productAll)

# BenefitInfo={}
# BenefitInfo1={}
# BenefitInfo2={}
# BenefitInfo3={}
# BenefitInfo4={}
# BenefitInfo5={}

# for col in range(1,sheet3.max_column+1):
#         BenefitInfo[sheet3.cell(3, col).value] = sheet3.cell(4, col).value
#         BenefitInfo1[sheet3.cell(3, col).value] = sheet3.cell(5, col).value
#         BenefitInfo2[sheet3.cell(3, col).value] = sheet3.cell(6, col).value
#         BenefitInfo3[sheet3.cell(3, col).value] = sheet3.cell(7, col).value
#         BenefitInfo4[sheet3.cell(3, col).value] = sheet3.cell(8, col).value
#         BenefitInfo5[sheet3.cell(3, col).value] = sheet3.cell(9, col).value
# print(BenefitInfo)

# Agent={}
# for col in range(1,sheet4.max_column+1):
#         Agent[sheet4.cell(3, col).value] = sheet4.cell(4, col).value
# print(Agent)



# # ------------------------------------------------------------------------
# # data = [(1, 2, 3), (4, 5, 9)]
# # All_test_Data=[(personinfoAll,BenefitInfo,SaleReportAll,productAll),(personinfo1,BenefitInfo1,SaleReport1,product1)]
# All_test_Data=[
#         (personinfoAll,BenefitInfo,SaleReportAll,productAll),
#         (personinfo1,BenefitInfo1,SaleReport1,product1),
#         (personinfo2,BenefitInfo2,SaleReport2,product2),
#         (personinfo3,BenefitInfo3,SaleReport3,product3),
#         (personinfo4,BenefitInfo4,SaleReport4,product4),
#         (personinfo5,BenefitInfo5,SaleReport5,product5),
#         ]

# @pytest.fixture(scope="session",params=All_test_Data)
# def my_fixture(request):
#     return request.param


