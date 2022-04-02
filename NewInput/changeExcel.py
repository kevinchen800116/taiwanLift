from openpyxl import load_workbook

workbook = load_workbook(
    "C:\\Users\\user\\.jenkins\\workspace\\FileJob\\Data\\測試用.xlsx")

sheet = workbook["personinfoAll"]
sheet1 = workbook["SaleReportAll"]
sheet2 = workbook["productAll"]
sheet3 = workbook["BenefitInfo"]
sheet4 = workbook["Agent"]

test = []
for row in range(4, sheet.max_row+1):
    # 定義一個空字典
    dict = {}
    for col in range(1, sheet.max_column+1):
        # 迴圈將sheet內的資料一行一行寫入字典
        dict[sheet.cell(3, col).value] = sheet.cell(row, col).value
    # 將寫入資料後的字典，加進test串列
    test.append(dict)

# SaleReportAll
test1 = []
for row in range(4, sheet1.max_row+1):
    dict = {}
    for col in range(1, sheet1.max_column+1):
        dict[sheet1.cell(3, col).value] = sheet1.cell(row, col).value
    test1.append(dict)

# productAll
test2 = []
for row in range(4, sheet2.max_row+1):
    dict = {}
    for col in range(1, sheet2.max_column+1):
        dict[sheet2.cell(3, col).value] = sheet2.cell(row, col).value
    test2.append(dict)

# BenefitInfo
test3 = []
for row in range(4, sheet3.max_row+1):
    dict = {}
    for col in range(1, sheet3.max_column+1):
        dict[sheet3.cell(3, col).value] = sheet3.cell(row, col).value
    test3.append(dict)

# Agent
test4 = []
for row in range(4, sheet4.max_row+1):
    dict = {}
    for col in range(1, sheet4.max_column+1):
        dict[sheet4.cell(3, col).value] = sheet4.cell(row, col).value
    test4.append(dict)


All_test_Data = []
# 此處的range為控制每個頁籤內，row的列數。此處選擇test的頁籤內的全部資料，當作最大長度。
# ex:若agent頁籤的資料少於6筆資料，但其他頁籤有6筆資料，則test4會報錯。

for i in range(0, len(test)):
    All = (test[i], test1[i], test2[i], test3[i], test4[i])
    print("我是ALL："+str(All))
    All_test_Data.append(All)

print("全部資料："+str(All_test_Data))


# workbook.save(filename="NewInput_Data.xlsx")
workbook.save(
    "C:\\Users\\user\\Desktop\\PortableGit\\Zrealtest\\NewInput\\NewInput_Data.xlsx")
