import pytest
from openpyxl import load_workbook

workbook = load_workbook("Policy_number_Data.xlsx")

sheet = workbook["Policy_number"]



test=[]
for row in range(4,sheet.max_row+1):
        ## 定義一個空字典
        dict={}
        for col in range(1,sheet.max_column+1):
                ## 迴圈將sheet內的資料一行一行寫入字典
                dict[sheet.cell(3, col).value] = sheet.cell(row, col).value
        ## 將寫入資料後的字典，加進test串列
        test.append(dict)

All_test_Data=[]
# 此處的range為控制每個頁籤內，row的列數。此處選擇test的頁籤內的全部資料，當作最大長度。
# ex:若agent頁籤的資料少於6筆資料，但其他頁籤有6筆資料，則test4會報錯。 
for i in range(0,len(test)):
        All=(test[i])
        # print("我是ALL："+str(All))
        All_test_Data.append(All)

# Policy_number0={}
# Policy_number1={}
# Policy_number2={}
# Policy_number3={}

# for col in range(1,sheet.max_column+1):
#     Policy_number0[sheet.cell(3, col).value] = sheet.cell(4, col).value
#     Policy_number1[sheet.cell(3, col).value] = sheet.cell(5, col).value
#     Policy_number2[sheet.cell(3, col).value] = sheet.cell(6, col).value
#     Policy_number3[sheet.cell(3, col).value] = sheet.cell(7, col).value


# All_test_Data=[
#     (Policy_number0),
#     (Policy_number1),
#     (Policy_number2),
#     (Policy_number3),
#     ]
    
print(All_test_Data)

@pytest.fixture(scope="session",params=All_test_Data)
def my_fixture(request):
    return request.param